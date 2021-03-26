import io
import openpyxl

def load_brand(brand_file):
    brand_dict = {}
    with io.open(brand_file,"r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 2:continue
            brand_name,brand_id = line_list
            brand_dict[str(brand_name.strip())] = str(brand_id.strip())
    return brand_dict


def score_reg(similar_score):
    score_scope = ""
    if float(similar_score) >= 0.95:
        score_scope = 'score_95'
    elif float(similar_score) < 0.95 and float(similar_score) >= 0.94:
        score_scope = 'score_94'
    elif float(similar_score) < 0.94 and float(similar_score) >= 0.93:
        score_scope = 'score_93'
    elif float(similar_score) < 0.93:
        score_scope = 'score_low'
    return score_scope


def brand_stat(brand_file,sku_file):

    brand_dict = load_brand(brand_file)
    appoint_brand_sku_all_dict = {}
    appoint_brand_sku_dict = {}
    appoint_brand_sku_rev_dict = {}
    brand_top_sku_dict = {}

    with io.open(sku_file,"r",encoding="utf-8") as f1:
        for line in f1:
            if len(line) == 0:continue
            line_list = line.strip().split("\t")
            if len(line_list) != 18:continue
            tmall_brand_id = line_list[6]

            if line_list[14] == 'NULL':continue
            # print(brand_dict)
            if tmall_brand_id in brand_dict:
                if tmall_brand_id in appoint_brand_sku_dict:
                    appoint_brand_sku_dict[tmall_brand_id].append(line_list)
                else:
                    appoint_brand_sku_dict[tmall_brand_id] = [line_list]

    # print(appoint_brand_sku_dict['Moutai/茅台'])
    # print(appoint_brand_sku_dict)

    for brand_id,sku_lst in appoint_brand_sku_dict.items():
        sku_lst_rev = sorted(sku_lst, key=lambda x: float(x[9]), reverse=True)
        sku_lst_rev_top_ = sku_lst_rev[:100]
        brand_top_sku_dict[brand_id] = sku_lst_rev_top_
        # sku_lst_rev_top = sku_lst_rev
        # tmp_brand_dict["brand_id"] = brand_id
        # tmp_brand_dict["sku_lst_rev_top"] = sku_lst_rev_top
        appoint_brand_sku_all_dict[brand_id] = sku_lst_rev

    for brand_id,sku_lst_rev_top in appoint_brand_sku_all_dict.items():
        tmp_brand_dict = {}
        # print(sku_lst_rev_top)
        for sku_item in sku_lst_rev_top:
            score_scope = score_reg(sku_item[14])
            if score_scope == 'score_95':
                if 'score_95' in tmp_brand_dict:
                    tmp_brand_dict['score_95']['num_rate'] += 1
                    tmp_brand_dict['score_95']['gmv_rate'] += float(sku_item[9])
                else:
                    tmp_dict = {}
                    tmp_dict['num_rate'] = 1
                    tmp_dict['gmv_rate'] = float(sku_item[9])
                    tmp_brand_dict['score_95'] = tmp_dict

            elif score_scope == 'score_94':
                if 'score_94' in tmp_brand_dict:
                    tmp_brand_dict['score_94']['num_rate'] += 1
                    tmp_brand_dict['score_94']['gmv_rate'] += float(sku_item[9])
                else:
                    tmp_dict = {}
                    tmp_dict['num_rate'] = 1
                    tmp_dict['gmv_rate'] = float(sku_item[9])
                    tmp_brand_dict['score_94'] = tmp_dict

            elif score_scope == 'score_93':
                if 'score_93' in tmp_brand_dict:
                    tmp_brand_dict['score_93']['num_rate'] += 1
                    tmp_brand_dict['score_93']['gmv_rate'] += float(sku_item[9])
                else:
                    tmp_dict = {}
                    tmp_dict['num_rate'] = 1
                    tmp_dict['gmv_rate'] = float(sku_item[9])
                    tmp_brand_dict['score_93'] = tmp_dict

            else:
                if 'score_low' in tmp_brand_dict:
                    tmp_brand_dict['score_low']['num_rate'] += 1
                    tmp_brand_dict['score_low']['gmv_rate'] += float(sku_item[9])
                else:
                    tmp_dict = {}
                    tmp_dict['num_rate'] = 1
                    tmp_dict['gmv_rate'] = float(sku_item[9])
                    tmp_brand_dict['score_low'] = tmp_dict

        appoint_brand_sku_rev_dict[brand_id] = tmp_brand_dict

    return appoint_brand_sku_rev_dict


def writeExcel(path, value, sheet):
    '''
    :param sheet:sheet的名称
    :param path:文件的名字和路径
    :param value1: 写入的数据
    :return:
    '''
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = sheet

    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i + 1, column=j + 1, value=str(value[i][j]))

    book.save(path)
    print("写入数据成功！")


def addExcel(path, value, sheet):
    '''
	:param sheet:sheet的名称
    :param path:写入excel的路径
    :param value: 追加的数据
    :return:
    '''
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)
    wb.save(path)
    print("写入成功")


if __name__ == "__main__":
    brand_file = "./appoint_brand.txt"
    sku_file = "./4ji_shipin_data.txt"
    appoint_brand_sku_rev_dict = brand_stat(brand_file,sku_file)

    excel_list = []

    for brand_name,score_dict in appoint_brand_sku_rev_dict.items():
        score_95 = score_dict['score_95']

        if 'score_94' in score_dict:
            score_94 = score_dict['score_94']
        else:
            score_94 = {'num_rate': 0, 'gmv_rate': 0}

        if 'score_93' in score_dict:
            score_93 = score_dict['score_93']
        else:
            score_93 = {'num_rate':0,'gmv_rate':0}

        if 'score_low' in score_dict:
            score_low = score_dict['score_low']
        else:
            score_low = {'num_rate':0,'gmv_rate':0}

        all_num = score_95['num_rate'] + score_94['num_rate'] + score_93['num_rate'] + score_low['num_rate']
        all_gmv = score_95['gmv_rate'] + score_94['gmv_rate'] + score_93['gmv_rate'] + score_low['gmv_rate']
        excel_list.append([brand_name,score_95['num_rate'],score_94['num_rate'],score_93['num_rate'],score_low['num_rate'],\
                           score_95['gmv_rate'],score_94['gmv_rate'],score_93['gmv_rate'],score_low['gmv_rate']])

    excel_list.insert(0,['品牌','大于0.95数量','大于0.94数量','大于0.93数量','小于0.93数量',\
                         '大于0.95gmv','大于0.94gmv','大于0.93gmv','小于0.93gmv',])

    writeExcel("./四级类分数分布.xlsx",excel_list,"品牌分数分布")



    # print(appoint_brand_sku_rev_dict)

    # print(len(appoint_brand_sku_rev_dict))

    # brand_dict = load_brand(brand_file)
    # print(brand_dict)
    # s_list = [[2,3,4,5],[8,8,8,8],[3,4,5,6],[1,1,1,1],[5,6,7,8]]
    # s_list_rev = sorted(s_list,key = lambda x:x[2],reverse=True)
    # print(s_list_rev)